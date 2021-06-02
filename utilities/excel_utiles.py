import openpyxl as xl


# get row count of the excel sheet
def get_row_count(file, sheet_name):
    workbook = xl.load_workbook(file)
    sheet_name = workbook[sheet_name]
    row_count = sheet_name.max_row
    return row_count


# get column count of the excel sheet
def get_column_count(file, sheet_name):
    workbook = xl.load_workbook(file)
    sheet_name = workbook[sheet_name]
    column_count = sheet_name.max_column
    return column_count


# read data from excel file
def read_data(file, sheet_name, row_number, column_number):
    workbook = xl.load_workbook(file)
    sheet_name = workbook[sheet_name]
    cell_value = sheet_name.cell(row_number, column_number).value
    return cell_value


# write data into excel file
def write_data(file, sheet_name, row_number, column_number, data):
    workbook = xl.load_workbook(file)
    sheet_name = workbook[sheet_name]
    cell_value = sheet_name.cell(row_number, column_number).value
    cell_value = data
    workbook.save(file)
    workbook.close()
